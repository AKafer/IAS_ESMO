import json
from datetime import datetime

from openpyxl import Workbook

from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.views import View
import logging

from openpyxl.styles import Font, Alignment

from externals.esmo import esmo_client

logger = logging.getLogger("esmo")


TOKEN = 'eyJ0eXAiOiJKV1QiLCJhbGciOiJSUzI1NiIsImp0aSI6IjkwMGIyMTE0NWEzZjVjMDE0Nzg1MjdiZGQxN2RlOWIzZDNjOWZjOGEyODQ1OTk0NTI4ZGUzMjNjMmMzMjM4ZGRlM2M3ZTVmZmM3NTY3NWZjIn0.eyJhdWQiOiIxIiwianRpIjoiOTAwYjIxMTQ1YTNmNWMwMTQ3ODUyN2JkZDE3ZGU5YjNkM2M5ZmM4YTI4NDU5OTQ1MjhkZTMyM2MyYzMyMzhkZGUzYzdlNWZmYzc1Njc1ZmMiLCJpYXQiOjE3MTEyNzc5MjUsIm5iZiI6MTcxMTI3NzkyNSwiZXhwIjoxNzQyODEzOTI1LCJzdWIiOiIxNDAiLCJzY29wZXMiOltdfQ.LW3rw5VXrTsmYWKquAtkK_8UgOeifiji7t2N5Qf8LYx33JZN77-pDBH60zT_XFj0TXWQEwrg7eujFASZuYSybr7iBGhvnJF4fM5yiVvxDXRTBMBv6EY9jw-nFdyD9Qp-zUZRuQ1mfQkUWtEYOP0tuvV_VA3bUY2XAiJF61JagqlmRtpfbry-VPcbWFuZmtOmLq4IZe6x5mlgwLAme15ESTHs_Mq6IQrsF9nmSCXoCzORaKo6Eh_EnN21sSu_0cOmlR7yxQwUUlODs0i2KTVTQPDlesSZweqPUXGpR1pQyuUoG7u8rCNtOd2XZOAZt6zaUmjmeQyODCx5hyTKSdnliZtzc3qLVr2ZWgBm1pvI9hQm9E8JuerYAv--ls6xVIGLn3xS5YgeUIWgPAzCmH_vlukOh7pZ9YYd2TR8v8WyfNI4ISvZRkUOnF6EgFHfGJ7LIbtq69-MzXQnBoOZL2RbNUSFQeO0cCx4cehsePFBzf5o_dTEPqtO5UJmmJOMIvs7ElCUR1wtl6a-v62HbDQCqLuZv3rZO-Dyg1O2z2081fvKWH9LxpySyZFXtOMQTufqI2iysMGXTeiykw8IwQVh4jNF95TVE-3JlAbOkJoztN7CPpdYJNF0-UkU-8eYsZ7Ey7pdmwu7bJQCPu0rzH7lK0OANt5PO5Up7DNRmxIIJOM'

HEADERS = {
        'Authorization': f'Bearer {TOKEN}',
        'Content-Type': 'application/json',
        'Accept': 'application/json',
    }


class IndexView(View):
    async def get(self, request: HttpRequest) -> HttpResponse:
        return render(request, 'index.html')


class ApiTableView(View):
    async def get(self, request: HttpRequest):
        date = request.GET.get('date', '')
        if not date:
            return HttpResponse('Date is not provided', status=400)
        result = await esmo_client.get_examsessions(date)
        converted_result = json.dumps(result, default=str)
        return HttpResponse(converted_result, content_type="application/json")


def add_null_for_hour(date: datetime) -> str:
    if date:
        date = date.strftime('%Y-%m-%d %H:%M:%S')
        date_str, time_str = date.split(' ')
        if len(time_str) < 8:
            date = date_str + ' ' + '0' + time_str
    return date


class ApiFileView(View):
    async def get(self, request: HttpRequest):
        date = request.GET.get('date', '')
        if not date:
            return HttpResponse('Date is not provided', status=400)

        # create workbook
        wb = Workbook()
        sheet = wb.active

        # stylize header row
        c1 = sheet.cell(row=1, column=1)
        c1.value = "№/№"
        c1.font = Font(bold=True)
        c1.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['A'].width = 10

        c2 = sheet.cell(row=1, column=2)
        c2.value = "Номер работника"
        c2.font = Font(bold=True)
        c2.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['B'].width = 20

        c3 = sheet.cell(row=1, column=3)
        c3.value = "Полное имя"
        c3.font = Font(bold=True)
        c3.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['C'].width = 35

        c4 = sheet.cell(row=1, column=4)
        c4.value = "Тест тип 1"
        c4.font = Font(bold=True)
        c4.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['D'].width = 15

        c5 = sheet.cell(row=1, column=5)
        c5.value = "Дата"
        c5.font = Font(bold=True)
        c5.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['E'].width = 25

        c6 = sheet.cell(row=1, column=6)
        c6.value = "Тест тип 2"
        c6.font = Font(bold=True)
        c6.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['F'].width = 15

        c7 = sheet.cell(row=1, column=7)
        c7.value = "Дата"
        c7.font = Font(bold=True)
        c7.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['G'].width = 25

        c8 = sheet.cell(row=1, column=8)
        c8.value = "Продолжительность"
        c8.font = Font(bold=True)
        c8.alignment = Alignment(horizontal="center")
        sheet.column_dimensions['H'].width = 25

        # export data to Excel
        result = await esmo_client.get_examsessions(date)
        for idx, item in enumerate(result):
            sheet.append([
                idx + 1,
                item['number'],
                item['name'],
                item['type_1'],
                add_null_for_hour(item['date_type_1']),
                item['type_2'],
                add_null_for_hour(item['date_type_2']),
                item['duration'],
            ])
            for i in range(1, 9):
                cell = sheet.cell(row=idx + 2, column=i)
                cell.alignment = Alignment(horizontal="right")

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Exams_data_{date}.xlsx"'.format(date=date)

        wb.save(response)

        return response
