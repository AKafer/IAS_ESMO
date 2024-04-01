from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def get_time(diff: timedelta) -> str:
    h = diff.seconds // 3600
    if h < 10:
        h = f"0{h}"
    h = f"{h} ч.,"
    m = diff.seconds // 60 % 60
    if m < 10:
        m = f"0{m}"
    m = f" {m} м."
    return h + m


def get_empl_dict(employees: list[dict]) -> dict:
    return {empl["uuid"]: empl for empl in employees}


def exam_handler(exam_list: list[dict], employee_dict: dict) -> list[dict]:
    res = {}
    for exam in exam_list:
        if exam["employee_uuid"] not in res:
            res[exam["employee_uuid"]] = {
                "number": exam["employee_uuid"],
                "name": "Не найдено",
                "type_1": '',
                "date_type_1": None,
                "type_2": '',
                "date_type_2": None,
                "duration": "00 ч., 00 м.",
            }
        if exam["type"] == 1:
            res[exam["employee_uuid"]]["type_1"] = 'X'
            res[exam["employee_uuid"]]["date_type_1"] = datetime.fromtimestamp(exam["date"])
        if exam["type"] == 2:
            res[exam["employee_uuid"]]["type_2"] = 'X'
            res[exam["employee_uuid"]]["date_type_2"] = datetime.fromtimestamp(exam["date"])
        if (
            res[exam["employee_uuid"]]["type_1"]
            and res[exam["employee_uuid"]]["type_2"]
            and res[exam["employee_uuid"]]["date_type_1"] < res[exam["employee_uuid"]]["date_type_2"]
        ):
            diff = res[exam["employee_uuid"]]["date_type_2"] - res[exam["employee_uuid"]]["date_type_1"]
            res[exam["employee_uuid"]]["duration"] = get_time(diff)
    for uuid, val in res.items():
        res[uuid]["name"] = employee_dict.get(uuid, {}).get("full_name", "Не найдено")
    res_list = list(res.values())
    return res_list


def add_null_for_hour(date: datetime) -> str:
    if date:
        date = date.strftime('%Y-%m-%d %H:%M:%S')
        date_str, time_str = date.split(' ')
        if len(time_str) < 8:
            date = date_str + ' ' + '0' + time_str
    return date


async def get_book(exams: list[dict]) -> Workbook:
    wb = Workbook()
    sheet = wb.active

    c1 = sheet.cell(row=1, column=1)
    c1.value = "№/№"
    c1.font = Font(bold=True)
    c1.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['A'].width = 10

    c2 = sheet.cell(row=1, column=2)
    c2.value = "Номер работника"
    c2.font = Font(bold=True)
    c2.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['B'].width = 20

    c3 = sheet.cell(row=1, column=3)
    c3.value = "Полное имя"
    c3.font = Font(bold=True)
    c3.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['C'].width = 35

    c4 = sheet.cell(row=1, column=4)
    c4.value = "Экзамен 1"
    c4.font = Font(bold=True)
    c4.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['D'].width = 10

    c5 = sheet.cell(row=1, column=5)
    c5.value = "Дата экзамена 1"
    c5.font = Font(bold=True)
    c5.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['E'].width = 20

    c6 = sheet.cell(row=1, column=6)
    c6.value = "Экзамен 2"
    c6.font = Font(bold=True)
    c6.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['F'].width = 10

    c7 = sheet.cell(row=1, column=7)
    c7.value = "Дата экзамена 2"
    c7.font = Font(bold=True)
    c7.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['G'].width = 20

    c8 = sheet.cell(row=1, column=8)
    c8.value = "Продолжительность"
    c8.font = Font(bold=True)
    c8.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['H'].width = 20

    for idx, item in enumerate(exams):
        sheet.append([
            idx + 1,
            item['number'],
            item['name'],
            item['type_1'],
            add_null_for_hour(item['date_type_1']),
            item['type_2'],
            add_null_for_hour(item['date_type_2']),
            item['duration'],
        ])
        for i in range(1, 9):
            cell = sheet.cell(row=idx + 2, column=i)
            cell.alignment = Alignment(horizontal="right")
    return wb
