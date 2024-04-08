from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def add_null_for_hour(date_lst: list[datetime]) -> str:
    new_lst = ""
    for date in date_lst:
        if date:
            date = date.strftime('%Y-%m-%d %H:%M:%S')
            date_str, time_str = date.split(' ')
            if len(time_str) < 8:
                date = date_str + ' ' + '0' + time_str
            new_lst += f"{date}, "
    return new_lst


def make_str(lst: list[str]) -> str:
    str_lst = ""
    for l in lst:
        str_lst += f"{l}, "
    return str_lst


async def get_book(exams: list[dict]) -> Workbook:
    wb = Workbook()
    sheet = wb.active

    c1 = sheet.cell(row=1, column=1)
    c1.value = "№/№"
    c1.font = Font(bold=True)
    c1.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['A'].width = 10

    c2 = sheet.cell(row=1, column=2)
    c2.value = "Номер"
    c2.font = Font(bold=True)
    c2.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['B'].width = 20

    c3 = sheet.cell(row=1, column=3)
    c3.value = "ФИО"
    c3.font = Font(bold=True)
    c3.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['C'].width = 40

    c4 = sheet.cell(row=1, column=4)
    c4.value = "Подразделение"
    c4.font = Font(bold=True)
    c4.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['D'].width = 30

    c5 = sheet.cell(row=1, column=5)
    c5.value = "Тест 1"
    c5.font = Font(bold=True)
    c5.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['E'].width = 20

    c6 = sheet.cell(row=1, column=6)
    c6.value = "Тест 2"
    c6.font = Font(bold=True)
    c6.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['F'].width = 20

    c7 = sheet.cell(row=1, column=7)
    c7.value = "Время работы"
    c7.font = Font(bold=True)
    c7.alignment = Alignment(horizontal="center")
    sheet.column_dimensions['G'].width = 30
    for idx, item in enumerate(exams):
        sheet.cell(idx + 2, 1).value = idx + 1
        sheet.cell(idx + 2, 2).value = item['number']
        sheet.cell(idx + 2, 3).value = item['name']
        sheet.cell(idx + 2, 4).value = item['division']
        sheet.cell(idx + 2, 5).value = add_null_for_hour(item['type_1'])
        sheet.cell(idx + 2, 6).value = add_null_for_hour(item['type_2'])
        sheet.cell(idx + 2, 7).value = make_str(item['duration'])
        for i in range(1, 8):
            cell = sheet.cell(row=idx + 2, column=i)
            if i in (5, 6):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")

    return wb
